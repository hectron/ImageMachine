import csv
import openpyxl
from collections import OrderedDict

def extract_listings(file_path):
    """
    This method extracts listings from a file and exports them to memory.
    Currently it only supports tab-delimited text files (.csv or .txt) and
    excel (.xlsx) files.
    """
    file_type = file_path.split('.')[-1].lower()

    if file_type == 'csv' or file_type == 'txt':
        return extract_csv_listings(file_path)
    elif file_type == 'xls' or file_type == 'xlsx':
        return extract_xlsx_listings(file_path)
    else:
        print("[ERROR] Unsupported file type: {0}".format(file_type))
        return []

def extract_csv_listings(file_path):
    """
    Extracts the listings as a dictionary from a tab-delimited file (.csv,
    .txt) file. It returns an array of dictionaries.
    """
    listings = []

    with open(file_path, 'r') as f:
        csv_reader = csv.DictReader(f)

        for row in csv_reader:
            listings.append(row)

    return listings

def extract_xlsx_listings(file_path):
    """
    Extracts the listings as a dictionary from a .xslx file. It returns an
    array of dictionaries.
    """
    listings = []

    # Do not include formulas, get the calculated data
    workbook = openpyxl.load_workbook(filename = file_path, read_only=True, data_only=True)

    sheet_names = workbook.get_sheet_names()

    for sheet in sheet_names:
        s = workbook[sheet]
        i = 0
        headers = []
        listings = []

        for row in s.rows:
            real_row = OrderedDict()

            if i == 0:
                for cell in row:
                    headers.append(cell.value)
            else:
                cell_index = 0

                for cell in row:
                    real_row[headers[cell_index]] = cell.value
                    cell_index += 1

            i += 1

            listings.append(real_row)

    return listings

def unique_listings(*list_of_listings):
    unique_listings = []

    for listings in list_of_listings:
        for listing in listings:
            if listing not in unique_listings:
                unique_listings.append(listing)

    return unique_listings


def write_listings(new_file_path, listings):
    workbook = openpyxl.Workbook()

    sheet = workbook.active
    sheet.title = 'cleaned listings'

    headers = listings[0].keys()

    cell_index = 0
    for cell in headers:
        sheet.cell(column=cell_index, row=0, value=cell)
        cell_index += 1

    row_index = 1
    for listing in listings:
        cell_index = 0
        for key in headers:
            sheet.cell(column=cell_index, row=row_index, value=listing[key])
            cell_index += 1

        row_index += 1

    workbook.save(filename = new_file_path)
